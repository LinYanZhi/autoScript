import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from typing import List, Tuple, Dict, Union, Optional
import openpyxl.worksheet
import openpyxl.worksheet.worksheet


class ExcelTool:
    """Excel 文件操作工具类（支持读取/写入单元格值及样式）"""
    
    @staticmethod
    def load_workbook(file_path: str) -> openpyxl.Workbook:
        """加载 Excel 文件"""
        return openpyxl.load_workbook(file_path)
    
    @staticmethod
    def save_workbook(workbook: openpyxl.Workbook, file_path: str) -> None:
        """保存 Excel 文件"""
        workbook.save(file_path)
    
    @staticmethod
    def select_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """选择工作表（Sheet）"""
        return workbook[sheet_name]
    
    @staticmethod
    def read_cell(sheet: openpyxl.worksheet.worksheet.Worksheet,
                  row: int, col: int) -> Union[str, int, float, None]:
        """读取单元格的值 (行和列从1开始计数)"""
        return sheet.cell(row=row, column=col).value
    
    @staticmethod
    def write_cell(sheet: openpyxl.worksheet.worksheet.Worksheet,
                   row: int, col: int, value: Union[str, int, float]) -> None:
        """写入单元格的值"""
        sheet.cell(row=row, column=col, value=value)
    
    @staticmethod
    def read_column(sheet: openpyxl.worksheet.worksheet.Worksheet,
                    col: int, min_row: int = 1, max_row: Optional[int] = None) -> List:
        """读取整列数据 (自动检测列尾)"""
        max_row = max_row or sheet.max_row
        return [sheet.cell(row=r, column=col).value for r in range(min_row, max_row + 1)]
    
    @staticmethod
    def read_row(sheet: openpyxl.worksheet.worksheet.Worksheet,
                 row: int, min_col: int = 1, max_col: Optional[int] = None) -> List:
        """读取整行数据 (自动检测列尾)"""
        max_col = max_col or sheet.max_column
        return [sheet.cell(row=row, column=c).value for c in range(min_col, max_col + 1)]


if __name__ == '__main__':
    path = r"C:\Users\Administrator\Desktop\export_data.xlsx"
    
    # 加载 Excel 文件
    wb = ExcelTool.load_workbook(path)
    
    # 选择 sheet 页
    sheet = ExcelTool.select_sheet(wb, "Sheet1")
    
    # 读取单元格
    # cell_value = ExcelTool.read_cell(sheet, row=2, col=1)
    # print(f"A2 单元格值: {cell_value}")
    
    # 写入单元格
    # ExcelTool.write_cell(sheet, row=1, col=2, value="新数据")
    
    # 读取整列
    # column_data = ExcelTool.read_column(sheet, col=3)  # 读取C列
    # print(f"C列数据: {column_data}")
    
    # 读取整行
    row_data = ExcelTool.read_row(sheet, row=1)  # 读取第2行
    print(f"第1行数据: {row_data}")
    
    # # 保存修改
    # ExcelTool.save_workbook(wb, "modified_example.xlsx")
